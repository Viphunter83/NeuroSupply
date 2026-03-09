
import asyncio
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional, List
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import io

from src.db.models import Order, OrderStatus, Restaurant, Product
from src.db.models.analytics import Anomalies
from src.schemas.order import OrderResponse
from src.services.data_loader.sheets_client import SheetsClient

logger = logging.getLogger(__name__)

class OrderService:
    def __init__(self, db: AsyncSession):
        self.db = db
        from src.services.calculation.engine_v2 import CalculationEngineV2
        self.engine = CalculationEngineV2(db)

    async def export_order_to_excel(self, order_id: UUID) -> io.BytesIO:
        """
        Generates an Excel file for the order.
        Columns: "Код", "Наименование", "Ед. изм.", "Кол-во (План)", "Кол-во (Факт)", "Комментарий"
        """
        stmt = select(Order).where(Order.id == order_id)
        result = await self.db.execute(stmt)
        order = result.scalar_one_or_none()
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        wb = Workbook()
        ws = wb.active
        ws.title = f"Order {str(order_id)[:8]}"
        
        # Headers
        headers = ["Код", "Наименование", "Ед. изм.", "Кол-во (План)", "Кол-во (Факт)", "Комментарий"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))

        # Data
        # items is a list of dicts.
        # Structure: product_id, product_name, unit, quantity (fact), predicted_usage (plan)
        
        for row_idx, item in enumerate(order.items, 2):
            # item is dict
            code = item.get('product_id', '') # Or iiko_id if available? Using internal ID for now or name
            name = item.get('product_name', 'Unknown')
            unit = item.get('unit', '')
            qty_plan = item.get('predicted_usage', 0)
            qty_fact = item.get('quantity', 0)
            
            ws.cell(row=row_idx, column=1, value=str(code))
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=str(unit))
            ws.cell(row=row_idx, column=4, value=qty_plan)
            ws.cell(row=row_idx, column=5, value=qty_fact)
            ws.cell(row=row_idx, column=6, value=item.get('comment', '')) # Comment
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def get_active_order(self, restaurant_id: UUID) -> Optional[Order]:
        """
        Fetch the latest active order (DRAFT or VERIFIED_BY_COOK) 
        for a specific restaurant created TODAY.
        """
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        stmt = (
            select(Order)
            .where(
                Order.restaurant_id == restaurant_id,
                Order.status.in_([OrderStatus.DRAFT, OrderStatus.VERIFIED_BY_COOK]),
                Order.created_at >= today_start
            )
            .order_by(Order.created_at.desc())
            .limit(1)
        )
        result = await self.db.execute(stmt)
        return result.scalar_one_or_none()

    async def confirm_order(self, order_id: UUID) -> Order:
        """
        Confirm an order by changing its status to VERIFIED_BY_COOK.
        """
        stmt = select(Order).where(Order.id == order_id)
        result = await self.db.execute(stmt)
        order = result.scalar_one_or_none()

        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        if order.status != OrderStatus.DRAFT:
             # Depending on business logic, maybe allow re-confirming? 
             # For now, let's just log warning or allow it.
             logger.warning(f"Order {order_id} is already in state {order.status}")

        order.status = OrderStatus.VERIFIED_BY_COOK
        await self.db.commit()
        await self.db.refresh(order)
        
        logger.info(f"Order {order_id} confirmed (VERIFIED_BY_COOK).")
        return order

    async def approve_order(self, order_id: UUID) -> Order:
        """
        Approve an order by changing its status to APPROVED_BY_MANAGER.
        """
        stmt = select(Order).where(Order.id == order_id)
        result = await self.db.execute(stmt)
        order = result.scalar_one_or_none()

        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

        # Validation: only Verified orders can be approved
        if order.status != OrderStatus.VERIFIED_BY_COOK:
             logger.warning(f"Order {order_id} is in state {order.status}, expecting VERIFIED_BY_COOK")
             if order.status == OrderStatus.APPROVED_BY_MANAGER:
                 return order
             raise HTTPException(status_code=400, detail=f"Cannot approve order in {order.status} state")

        order.status = OrderStatus.APPROVED_BY_MANAGER
        await self.db.commit()
        await self.db.refresh(order)
        
        logger.info(f"Order {order_id} approved (APPROVED_BY_MANAGER).")
        return order


    async def _export_draft_to_sheet(self, restaurant_id: UUID, items: List[dict]):
        """Helper to export draft to sheet safely."""
        try:
            stmt = select(Restaurant).where(Restaurant.id == restaurant_id)
            result = await self.db.execute(stmt)
            restaurant = result.scalar_one_or_none()
            
            if restaurant and restaurant.spreadsheet_id:
                client = SheetsClient(restaurant.spreadsheet_id)
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, client.write_draft_order, items)
                logger.info(f"Exported draft order to sheet for {restaurant.name}")
            else:
                logger.warning(f"No spreadsheet_id for restaurant {restaurant_id}, skipping export.")
        except Exception as e:
            logger.error(f"Failed to export draft to sheet: {e}")

    async def generate_draft_order(self, restaurant_id: UUID, sales_plan_rub: float) -> Order:
        """
        Calculates needs and creates a new DRAFT order.
        Fetches dynamic settings (Safety Stock, Transit) from Google Sheets first.
        """
        # 0. Fetch Settings from Sheets (if possible)
        # We need restaurant object first
        stmt = select(Restaurant).where(Restaurant.id == restaurant_id)
        result = await self.db.execute(stmt)
        restaurant = result.scalar_one_or_none()
        
        calc_settings = {}
        if restaurant and restaurant.spreadsheet_id:
            try:
                client = SheetsClient(restaurant.spreadsheet_id)
                loop = asyncio.get_running_loop()
                # Run in executor to avoid blocking main thread
                calc_settings = await loop.run_in_executor(None, client.fetch_settings)
                logger.info(f"Fetched settings for {restaurant.name}: {calc_settings}")
            except Exception as e:
                logger.warning(f"Could not fetch settings from sheet, using defaults: {e}")

        # 1. Calculate Needs (using V2 Logic)
        # Now returns tuple: (items, dish_breakdown)
        # Pass settings if active
        # settings dict keys: 'safety_stock', 'days_in_transit'
        ss = calc_settings.get("safety_stock")
        dit = calc_settings.get("days_in_transit", 0)
        
        items, dish_breakdown = await self.engine.calculate_needs(
            restaurant_id, 
            sales_plan_rub,
            safety_stock=ss,
            days_in_transit=dit
        )
        
        # 2. Create Order in DB
        new_order = Order(
            restaurant_id=restaurant_id,
            status=OrderStatus.DRAFT,
            items=items
        )
        self.db.add(new_order)
        await self.db.commit()
        await self.db.refresh(new_order)
        
        # 3. Export to Sheets (Feedback Loop)
        # 3.1 Export Draft Order (Tab 4)
        # We need to await these exports or fire-and-forget?
        # Awaiting ensures user sees it immediately.
        await self._export_draft_to_sheet(restaurant_id, items)
        
        # 3.2 Export Dish Calculation (Tab 2a) - NEW
        await self._export_dish_calc_to_sheet(restaurant_id, dish_breakdown)
        
        return new_order
        
    async def _export_dish_calc_to_sheet(self, restaurant_id: UUID, dishes: List[dict]):
        try:
            stmt = select(Restaurant).where(Restaurant.id == restaurant_id)
            result = await self.db.execute(stmt)
            restaurant = result.scalar_one_or_none()
            
            if restaurant and restaurant.spreadsheet_id:
                client = SheetsClient(restaurant.spreadsheet_id)
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, client.write_dish_calculation, dishes)
                logger.info(f"Exported dish calculation for {restaurant.name}")
        except Exception as e:
            logger.error(f"Failed to export dish calc: {e}")

    async def update_order_items(self, order_id: UUID, new_items: List[dict]) -> Order:
        """
        Updates order items and logs anomalies.
        """
        stmt = select(Order).where(Order.id == order_id)
        result = await self.db.execute(stmt)
        order = result.scalar_one_or_none()
        
        if not order:
             raise HTTPException(status_code=404, detail="Order not found")
             
        if order.status != OrderStatus.DRAFT:
             raise HTTPException(status_code=400, detail="Cannot update confirmed order")

        # Logic to detect anomalies (compare old vs new)
        
        # Pre-fetch products for all items to get package_size
        p_ids = [uuid.UUID(item.get('product_id')) for item in new_items if item.get('product_id')]
        p_stmt = select(Product).where(Product.id.in_(p_ids))
        p_res = await self.db.execute(p_stmt)
        p_map = {p.id: p for p in p_res.scalars().all()}

        # Map old items for comparison
        old_map = {item.get('product_id'): item for item in (order.items or [])}

        for new_item in new_items:
            p_id_str = new_item.get('product_id')
            p_id = uuid.UUID(p_id_str) if p_id_str else None
            new_qty = float(new_item.get('quantity', 0))
            old_item = old_map.get(p_id_str)
            
            # Recalculate quantity_kg based on product package_size
            product = p_map.get(p_id)
            if product:
                pkg_size = float(product.package_size) if product.package_size else 0.0
                new_item['quantity_kg'] = round(new_qty * pkg_size if pkg_size > 0 else new_qty, 4)

            if old_item:
                old_qty = float(old_item.get('quantity', 0))
                if abs(new_qty - old_qty) > 0.01:
                    # Anomaly detected
                    # Check if reason provided in item
                    reason = new_item.get('reason', "Manual update via API")
                    
                    if p_id in p_map:
                        anomaly = Anomalies(
                            order_id=order.id,
                            product_id=p_id,
                            auto_qty=old_qty,
                            manual_qty=new_qty,
                            reason=reason
                        )
                        self.db.add(anomaly)
            else:
                 # New item added?
                 # Treat as anomaly from 0 to New Qty
                 reason = new_item.get('reason', "Manual addition via API")
                 if p_id in p_map:
                     anomaly = Anomalies(
                         order_id=order.id,
                         product_id=p_id,
                         auto_qty=0.0,
                         manual_qty=new_qty,
                         reason=reason
                     )
                     self.db.add(anomaly)

        order.items = new_items
        await self.db.commit()
        await self.db.refresh(order)
        return order
